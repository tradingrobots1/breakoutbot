#!/usr/bin/env python3
"""
BREAKOUTBOT — Export a Excel

Vuelca outcomes.json a un .xlsx sencillo y legible. Se puede re-ejecutar
en cualquier momento para tener una foto actualizada.

Uso: python3 export_outcomes_excel.py
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

OUTCOMES_FILE = Path(__file__).parent / "outcomes.json"
OUT_FILE = Path(__file__).parent / "docs" / "Breakoutbot_Outcomes.xlsx"

ACCOUNT_CONFIG_FILE = Path(__file__).parent / "account_config.json"
_ACCOUNT_CONFIG_DEFAULTS = {"starting_capital": 2500.0, "risk_per_trade_pct": 0.5}


def _load_account_config() -> dict:
    """Misma fuente que breakout_scanner.py (account_config.json) — evita
    tener el capital inicial hardcodeado y desincronizado en cada script."""
    if ACCOUNT_CONFIG_FILE.exists():
        try:
            data = json.loads(ACCOUNT_CONFIG_FILE.read_text(encoding="utf-8"))
            return {**_ACCOUNT_CONFIG_DEFAULTS, **data}
        except Exception:
            pass
    return dict(_ACCOUNT_CONFIG_DEFAULTS)


STARTING_CAPITAL = _load_account_config()["starting_capital"]


def fmt_pct(v):
    if v is None:
        return ""
    if v == "N/A":
        return "N/A"
    return round(v, 1)


def get_current_equity(outcomes):
    equity = STARTING_CAPITAL
    for o in outcomes:
        if o.get("resolved") and o.get("position_value"):
            close_pct = o["checkpoints"].get("close")
            if isinstance(close_pct, (int, float)):
                equity += o["position_value"] * close_pct / 100
    return equity


def main():
    outcomes = json.loads(OUTCOMES_FILE.read_text(encoding="utf-8")) if OUTCOMES_FILE.exists() else []

    wb = Workbook()
    ws = wb.active
    ws.title = "Outcomes"

    headers = ["Ticker", "Fecha/hora entrada", "Precio entrada", "Acciones", "Valor posición",
               "Riesgo $", "ORB15", "PDH (intacto)", "Stop-loss", "% hoy al detectar",
               "m15 (%)", "m30 (%)", "h1 (%)", "h2 (%)", "close (%)", "P&L $ (close)",
               "Parada por stop", "Resuelto"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for o in sorted(outcomes, key=lambda x: x["entry_datetime"]):
        cps = o["checkpoints"]
        close_pct = cps.get("close")
        dollar_pnl = (round(o["position_value"] * close_pct / 100, 2)
                      if o.get("position_value") and isinstance(close_pct, (int, float)) else "")
        row = [
            o["ticker"],
            o["entry_datetime"][:16].replace("T", " "),
            round(o["entry_price"], 4),
            o.get("shares", ""),
            o.get("position_value", ""),
            o.get("risk_amount", ""),
            round(o["orb15"], 4) if o.get("orb15") else "",
            round(o["pdh"], 4) if o.get("pdh") else "",
            round(o["stop_price"], 4) if o.get("stop_price") else "",
            o.get("pct_change_seen", ""),
            fmt_pct(cps.get("m15")),
            fmt_pct(cps.get("m30")),
            fmt_pct(cps.get("h1")),
            fmt_pct(cps.get("h2")),
            fmt_pct(close_pct),
            dollar_pnl,
            "Sí" if o.get("stopped_out") else "No",
            "Sí" if o["resolved"] else "No",
        ]
        ws.append(row)

    # colorear positivos en verde, negativos en rojo (retorno long: + = gana)
    green = Font(color="1e7d34")
    red = Font(color="c0392b")
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=16):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.font = green if cell.value > 0 else red

    widths = [10, 17, 13, 9, 13, 10, 10, 13, 11, 15, 9, 9, 9, 9, 9, 12, 11, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # segunda hoja: resumen de cuenta
    ws2 = wb.create_sheet("Resumen")
    equity = get_current_equity(outcomes)
    rows = [
        ("Capital inicial", STARTING_CAPITAL),
        ("Equity actual", round(equity, 2)),
        ("P&L acumulado", round(equity - STARTING_CAPITAL, 2)),
        ("P&L acumulado (%)", round((equity - STARTING_CAPITAL) / STARTING_CAPITAL * 100, 2)),
        ("Entradas totales", len(outcomes)),
        ("Entradas resueltas", sum(1 for o in outcomes if o.get("resolved"))),
        ("Cerradas por stop-loss", sum(1 for o in outcomes if o.get("stopped_out"))),
    ]
    for label, value in rows:
        ws2.append([label, value])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    for r in range(1, len(rows) + 1):
        ws2.cell(row=r, column=1).font = Font(bold=True)

    OUT_FILE.parent.mkdir(exist_ok=True)
    wb.save(OUT_FILE)
    print(f"Excel generado: {OUT_FILE} ({len(outcomes)} filas) — equity actual ${equity:,.2f}")


if __name__ == "__main__":
    main()
